"""
Import historical picks from pivoted XLSX sheets.

Expected sheet format (per week):
- Sheet name like: "picks wk 6" (case-insensitive, spaces ok)
- Row (N-1): player names across columns (A is team abbr header)
- Row N    : player numbers (1..200) across columns (text or number)
- Then pairs of rows: AWAY team abbr on row R, HOME team abbr on row R+1
  e.g., "PHI" then "NYG" meaning PHI @ NYG
- A row starting with "DIST" marks end of data
- A "Consensus" column may exist; it will be ignored

We:
- Detect header rows robustly (numbers as text OK)
- Create/rename players (fake emails if needed)
- Map (week, home_abbr, away_abbr) to game_id
- Upsert picks
"""

# pylint: disable=line-too-long, too-many-locals, too-many-branches, broad-exception-caught

from __future__ import annotations
from typing import Any, Optional, List, Tuple, Dict
from dataclasses import dataclass
import logging
import os

from openpyxl import load_workbook

# ---------------------------
# Logging setup
# ---------------------------
LOG_LEVEL = os.getenv("PICKS_IMPORT_LOG_LEVEL", "INFO").upper()
logging.basicConfig(level=getattr(logging, LOG_LEVEL, logging.INFO),
                    format="%(levelname)s %(message)s")
log = logging.getLogger("picks_import")


# ---------------------------
# Spreadsheet → ESPN aliases
# (extend as needed)
# ---------------------------
TEAM_ALIASES: Dict[str, str] = {
    # Common mismatches
    "AZ": "ARI",
    "ARZ": "ARI",
    "PHX": "ARI",   # ancient
    "WAS": "WSH",
    "WASH": "WSH",
    "JAC": "JAX",
    # Legacy relocations
    "OAK": "LV",
    "SD": "LAC",
    "STL": "LAR",
    # Ambiguities (keep conservative)
    # "LA": "LAR",  # <- leave commented; ambiguous between LAR/LAC
}

def _norm_team(label: str) -> str:
    """Normalize a spreadsheet team label to ESPN/DB canonical code."""
    key = (label or "").strip().upper()
    return TEAM_ALIASES.get(key, key)


@dataclass
class _PlayerCol:
    col_idx: int
    pigeon_number: int
    pigeon_name: str


def _is_blank(v: Any) -> bool:
    if v is None:
        return True
    if isinstance(v, str) and v.strip() == "":
        return True
    return False


def _to_int_safe(v: Any) -> Optional[int]:
    """Parse Excel cell to int; accepts strings like ' 7 ' or '+7' and floats like 7.0."""
    if _is_blank(v):
        return None
    s = str(v).strip().replace("+", "")
    try:
        f = float(s)
        i = int(f)
        return i if abs(f - i) < 1e-9 else None
    except Exception:  # noqa: BLE001
        return None


def _detect_header(ws) -> Tuple[int, int, int, int]:
    """
    Return (name_row_idx, number_row_idx, first_player_col, last_player_col).
    Detect row with many ints as the player-number row; names are row-1.
    """
    max_ints = 0
    number_row = None
    first_col = None
    last_col = None

    for row in range(2, 15):  # look in first few rows
        ints_here = []
        for col in range(1, ws.max_column + 1):
            val = _to_int_safe(ws.cell(row=row, column=col).value)
            if val is not None:
                ints_here.append(col)
        if len(ints_here) > max_ints and len(ints_here) >= 10:
            max_ints = len(ints_here)
            number_row = row
            first_col = min(ints_here)
            last_col = max(ints_here)

    if number_row is None or first_col is None or last_col is None:
        raise RuntimeError("Could not detect player number row / columns.")

    name_row = number_row - 1
    return name_row, number_row, first_col, last_col


def _collect_players(ws, name_row: int, number_row: int, c1: int, c2: int) -> List[_PlayerCol]:
    players: List[_PlayerCol] = []
    for col in range(c1, c2 + 1):
        nm = ws.cell(row=name_row, column=col).value
        pn = _to_int_safe(ws.cell(row=number_row, column=col).value)
        if pn is None:
            continue
        name = (str(nm).strip() if nm is not None else f"Pigeon {pn}") or f"Pigeon {pn}"
        players.append(_PlayerCol(col_idx=col, pigeon_number=pn, pigeon_name=name))
    return players


def _iter_team_rows(ws, start_row: int, *, label_col: int = 1) -> list[str]:
    """
    Return a list of team labels starting at start_row (normalized).
    - label_col: 1-based column index holding the team label (e.g., 3 for column 'C').
    Stops when encountering 'DIST' or 'CONSENSUS'.
    Skips fully blank rows.
    """
    labels: list[str] = []
    max_r = ws.max_row or 0

    for r in range(start_row, max_r + 1):
        v = ws.cell(row=r, column=label_col).value
        if v is None or str(v).strip() == "":
            continue
        label = str(v).strip().upper()
        if label in {"DIST", "CONSENSUS"}:
            break
        labels.append(_norm_team(label))  # <-- normalize here
    return labels


def _find_game_id(cur, week: int, t1: str, t2: str) -> Optional[Tuple[int, str, str]]:
    """
    Attempt to find the game for (t1, t2) in any order.
    Returns (game_id, home_abbr, away_abbr) if found.
    """
    cur.execute(
        """
        SELECT game_id, home_abbr, away_abbr
        FROM games
        WHERE week_number = %s
          AND (
                (home_abbr = %s AND away_abbr = %s) OR
                (home_abbr = %s AND away_abbr = %s)
              )
        """,
        (week, t1, t2, t2, t1),
    )
    row = cur.fetchone()
    if not row:
        return None
    return int(row[0]), row[1], row[2]


def _upsert_player(cur, pigeon_number: int, pigeon_name: str) -> None:
    dummy_email = f"p{pigeon_number}@example.invalid"
    cur.execute(
        """
        INSERT INTO players (pigeon_number, pigeon_name, email, password_hash)
        VALUES (%s, %s, %s, '!')
        ON CONFLICT (pigeon_number) DO UPDATE SET pigeon_name = EXCLUDED.pigeon_name
        """,
        (pigeon_number, pigeon_name, dummy_email),
    )


def _upsert_pick(cur, pigeon_number: int, game_id: int, picked_home: bool, margin: int) -> None:
    cur.execute(
        """
        INSERT INTO picks (pigeon_number, game_id, picked_home, predicted_margin)
        VALUES (%s, %s, %s, %s)
        ON CONFLICT (pigeon_number, game_id)
        DO UPDATE SET picked_home = EXCLUDED.picked_home,
                      predicted_margin = EXCLUDED.predicted_margin,
                      created_at = now()
        """,
        (pigeon_number, game_id, picked_home, margin),
    )


def import_picks_pivot_xlsx(*, xlsx_path: str, conn, max_week: Optional[int] = None, only_week: Optional[int] = None) -> int:
    """
    Import a pivoted XLSX with one sheet per week (title like 'picks wk N').
    Uses alias normalization for team codes and logs detailed skip reasons.
    """
    wb = load_workbook(xlsx_path, data_only=True)
    processed = 0

    for ws in wb.worksheets:
        title = (ws.title or "").strip().lower()
        if not title.startswith("picks wk"):
            continue

        # parse week number from title
        try:
            week = int(title.split("wk", 1)[1].strip())
        except Exception:  # noqa: BLE001
            log.warning("[xlsx] Skipping sheet '%s' (week number parse failed).", ws.title)
            continue
        if only_week is not None and week != only_week:
            continue
        if max_week is not None and week > max_week:
            continue

        log.info("[xlsx] Processing '%s' as Week %d…", ws.title, week)

        # Header detection
        name_row, number_row, c1, c2 = _detect_header(ws)
        log.debug("[xlsx] Header: name_row=%s, number_row=%s, cols=%s..%s", name_row, number_row, c1, c2)
        players = _collect_players(ws, name_row, number_row, c1, c2)
        if not players:
            log.warning("[warn] No player columns with numbers detected on sheet '%s'", ws.title)
            continue

        cur = conn.cursor()
        for p in players:
            _upsert_player(cur, p.pigeon_number, p.pigeon_name)
        conn.commit()

        # Team rows (normalized)
        team_labels = _iter_team_rows(ws, number_row + 1, label_col=3)
        log.debug("[xlsx] First labels sample: %s", team_labels[:8])

        # Skip counters
        skips = {
            "odd_tail_row": 0,
            "no_game_match": 0,
            "both_blank": 0,
            "both_filled": 0,
            "non_int": 0,
        }
        sheet_count = 0

        i = 0
        while i < len(team_labels):
            if i + 1 >= len(team_labels):
                skips["odd_tail_row"] += 1
                break

            t_top = team_labels[i]     # usually AWAY (normalized)
            t_bot = team_labels[i + 1] # usually HOME (normalized)

            # Pre-scan: which players actually have numbers for this pair?
            players_with_numbers = []
            for p in players:
                v_top = ws.cell(row=number_row + 1 + i, column=p.col_idx).value
                v_bot = ws.cell(row=number_row + 1 + i + 1, column=p.col_idx).value
                top_int = _to_int_safe(v_top)
                bot_int = _to_int_safe(v_bot)
                if (top_int is not None) ^ (bot_int is not None):
                    players_with_numbers.append((p, top_int if top_int is not None else bot_int))

            # DB game match (in any order)
            cur = conn.cursor()
            found = _find_game_id(cur, week, t_top, t_bot)
            if not found:
                skips["no_game_match"] += 1
                if players_with_numbers:
                    # These picks could not be imported due to team-code mismatch
                    who = ", ".join([f"#{p.pigeon_number}({m})" for (p, m) in players_with_numbers[:12]])
                    more = "" if len(players_with_numbers) <= 12 else f" … +{len(players_with_numbers)-12} more"
                    log.warning(
                        "[miss] Week %d: no game match for pair '%s @ %s'; missed picks: %s%s",
                        week, t_top, t_bot, who, more
                    )
                else:
                    log.debug("[skip] Week %d: '%s @ %s' had no numbers; skipping.", week, t_top, t_bot)
                i += 2
                continue

            game_id, home_abbr, _ = found

            # For each player, write pick if exactly one of the two cells has an int
            for p in players:
                v_top = ws.cell(row=number_row + 1 + i, column=p.col_idx).value
                v_bot = ws.cell(row=number_row + 1 + i + 1, column=p.col_idx).value
                top_int = _to_int_safe(v_top)
                bot_int = _to_int_safe(v_bot)

                if top_int is None and bot_int is None:
                    skips["both_blank"] += 1
                    continue
                if top_int is not None and bot_int is not None:
                    skips["both_filled"] += 1
                    log.debug("[ambig] Week %d game_id=%s player #%s had numbers on both rows (%s & %s).",
                              week, game_id, p.pigeon_number, top_int, bot_int)
                    continue

                margin = top_int if top_int is not None else bot_int
                if margin is None:
                    skips["non_int"] += 1
                    log.debug("[nonint] Week %d game_id=%s player #%s had non-int value(s): top=%r bot=%r",
                              week, game_id, p.pigeon_number, v_top, v_bot)
                    continue

                team_with_pick = t_top if top_int is not None else t_bot
                picked_home = team_with_pick == home_abbr

                _upsert_pick(cur, p.pigeon_number, game_id, picked_home, margin)
                sheet_count += 1

            conn.commit()
            i += 2

        log.info("[xlsx] Week %d: processed %d player-pick cells. Skips=%s", week, sheet_count, skips)
        processed += sheet_count

    return processed
